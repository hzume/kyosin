import PySimpleGUI as sg
import openpyxl
from openpyxl.styles import PatternFill
import msoffcrypto
import tempfile
import os
from os.path import basename
import codecs
import re
import os
from src.views.popup import PasswordPopup, MeetingPopup
from src.views.error import *
from src.models.model import Tutor, Tutors, Class, Meeting
from src.settings.translate import *
from src.settings.config import *
from src.presenter.functions import *

class Presenter:
    def __init__(self, window: sg.Window):
        self.window = window
        self.tutors = Tutors()
        self.meetings = []

    # ---from Model---
    # 講師情報から名前を取得
    def get_tutor_names(self):
        tutor_names = []
        for tutor in self.tutors.values():
            if tutor.name != None and tutor.type != "運営":
                tutor_names.append(tutor.name)
        return tutor_names
    
    # ---to Model---
    # 講師情報全般を受け取り保存
    def init_tutors(self, params_list):
        for params in params_list:
            self.tutors.add(Tutor(**params))

    # 講師ごとの勤務時間を受け取り保存
    def set_worktime(self, params_list):
        for params in params_list:
            self.tutors.worktime_update(Class(**params))

    def update_config(self, values):
        with codecs.open("src/settings/config.py", "w", "utf-8") as f:
            f.write(
f"""default = {{
    'password': 1219,
    'tutor_path': '{values["tutor_path"]}',
    'admin_path': '{values["admin_path"]}',
    'template_path': '{values["template_path"]}',
    'year': {values["year"]},
    'output_folder': '{values["output_folder"]}',
    'gensen_path': '{values["gensen_path"]}'
}}
"""
            )
    
    # ミーティング情報を追加
    def add_meeting(self, meeting_day, meeting_length, participants):
        self.meetings.append(Meeting(meeting_day, meeting_length, participants))

    # ミーティング情報を講師に適用
    def set_meetings(self):
        for meeting in self.meetings:
            for participant in meeting.participants:
                self.tutors[participant].meeting[meeting.meeting_day] += meeting.meeting_length

    # ---from View---
    # ミーティング情報を読み取る
    def receive_meeting(self, tutor_names):
        popup = MeetingPopup(tutor_names)
        meeting_day, meeting_length, participants = popup.receive_meeting()
        del popup
        if meeting_day == "" or meeting_length == "":
            raise NotFilledError
        if not 1 <= int(meeting_day) <= 31:
            raise InvalidValue
        return meeting_day, meeting_length, participants

    # パスワードを読み取る
    def receive_password(self):
        popup = PasswordPopup()
        password = popup.receive_password()
        del popup
        if password == "":
            raise NotFilledError
        return int(password)
        
    # 講師情報を読取る
    def receive_tutors(self, tutor_path):
        wb = openpyxl.load_workbook(tutor_path)
        ws = wb.worksheets[0]
        params_list = []
        for row in ws.rows:
            if row[0].row == 1:
                keys = row
            else:
                params = {}
                for k, v in zip(keys, row):
                    params[ja_to_en[k.value]] = v.value
                params_list.append(params)
        return params_list
    
    # 授業の情報を読取る
    def receive_classes(self, admin_path, month, password, MAX_ROW = 100, MAX_COLUMN = 30):
        params_list = []
        with open(admin_path, "rb") as f, tempfile.TemporaryFile() as tf:
            ms_file = msoffcrypto.OfficeFile(f)
            ms_file.load_key(password=str(password))
            ms_file.decrypt(tf)
            wb = openpyxl.load_workbook(tf, data_only=True)
        trans_table = str.maketrans(zenkaku_to_hankaku)
        sheetnames = []

        for sheetname in wb.sheetnames:
            fixed_name = sheetname.translate(trans_table)
            idx = fixed_name.find("月")
            if idx == -1:
                continue
            if fixed_name[:idx] == str(month):
                sheetnames.append(sheetname)

        for sheetname in sheetnames:
            ws = wb[sheetname]
            flag = True
            for i in range(1, MAX_ROW):
                head_cell = ws.cell(row=i, column=2).value
                if (type(head_cell) == int) or ((head_cell not in class_times) and (head_cell not in class_marks)):
                    if type(head_cell) == int:
                        date = excel_date(head_cell)
                        if date.month != month:
                            flag = True # 月が異なるとTrue
                        else:
                            flag = False              
                        day = date.day - 1 # 0 ~ 30
                    if flag:
                        continue
                    table = ws[f"B{i}":f"BA{i+2}"]
                    # シート終了の判定
                    if table[2][0].value == None: 
                        break
                    class_time = class_times[table[2][0].value]
                    for j in range(1, MAX_COLUMN):
                        params = {}
                        name = table[0][j].value
                        class1 = table[1][j].value
                        class2 = table[2][j].value
                        trans_table = str.maketrans(zenkaku_to_hankaku)
                        if name != None and (class1 != None or class2 != None):
                            color = "ff8080"
                            table[0][j].fill = PatternFill(patternType="solid", fgColor=color, bgColor=color)
                            table[1][j].fill = PatternFill(patternType="solid", fgColor=color, bgColor=color)
                            table[2][j].fill = PatternFill(patternType="solid", fgColor=color, bgColor=color)
                            params["day"] = day
                            params["name"] = name
                            params["class_time"] = class_time
                            if (class1 != None) and ("事務" in class1):
                                params["officework"] = True
                                class1 = class1.translate(trans_table)
                                officework_time1 = re.findall(r"[0-9]+", class1)
                                if officework_time1:
                                    params["officework_time"] = int(officework_time1[0])
                            if (class2 != None) and ("事務" in class2):
                                params["officework"] = True
                                class2 = class2.translate(trans_table)
                                officework_time2 = re.findall(r"[0-9]+", class2)
                                params["officework_time"] = int(officework_time2[0])
                            
                            params_list.append(params)
            checksheet_path = os.path.dirname(admin_path)+f"/給与確認用{month}月分.xlsx"
            if os.path.exists(checksheet_path):
                os.remove(checksheet_path)
            wb.save(checksheet_path)
        return params_list
    
    # 源泉徴収税額を計算するExcel関数の生成
    def gensen(self, gensen_path):
        wb_gensen = openpyxl.load_workbook(gensen_path)
        ws_gensen = wb_gensen.worksheets[0]
        trans_table = str.maketrans(zenkaku_to_hankaku)
        rate_s = ws_gensen['L8'].value.translate(trans_table)
        rate = re.findall(r"[0-9]+\.?[0-9]*%", rate_s)
        if rate == None:
            raise Exception
        rate = float(rate[0][0:-1]) / 100.0
        amount_money = ws_gensen['B10'].value
        if amount_money == None:
            raise Exception
        s = "="
        s += f"IF(M2<{amount_money}, ROUNDDOWN(M2*{rate}, 0), "
        cnt = 1
        i = 0
        while amount_money < 200000:
            i += 1
            if ws_gensen[f'C{9+i}'].value == None:
                continue
            amount_money = ws_gensen[f'C{9+i}'].value
            cnt += 1
            tax = ws_gensen[f'L{9+i}'].value
            s += f"IF(M2<{amount_money}, {tax}, "
        s += "0" + ")"*cnt
        return s

    
    # ---to View---
    # テンプレートの更新、給与明細の作成
    def make_payslip(self, template_path, gensen_s, year, month, output_folder):
        for tutor in self.tutors.values():
            if tutor.type == "運営":
                continue
            if tutor.name == None:
                continue
            output_path = output_folder + "/" + tutor.fullname + f"{year}.xlsx"
            if not os.path.isfile(output_path):
                wb_template = openpyxl.load_workbook(template_path)
                wb_template.save(output_path)
            wb = openpyxl.load_workbook(output_path)

    # テンプレートの更新
            ws_template = wb.worksheets[0]
            for i in range(10,41):
                ws_template[f"M{i}"] = f"=IF(AND(COUNTA(H{i})=1,COUNTA(D{i}:H{i})=4),10, IF(AND(COUNTA(H{i})=1,COUNTA(D{i}:H{i})=5),40, 0))"
            ws_template["H5"].value = tutor.fullname
            ws_template["K41"].value = f"=ROUNDDOWN(SUM(K10:K40)/60*{tutor.pay_class},0)"
            ws_template["L41"].value = f"=ROUNDDOWN(SUM(L10:L40)/60*{tutor.pay_officework},0)"
            ws_template["M41"].value = f"=ROUNDDOWN(SUM(M10:M40)/240*{tutor.pay_officework}, 0)"
            ws_template["N41"].value = f"=ROUNDDOWN(SUM(N10:N40)/240*{tutor.pay_officework}, 0)"
            ws_template["P41"].value = f'=COUNTIF(P10:P40,"○")*{tutor.trans_fee}'
            
    # 給与明細作成
            ws_title = f"{year}年{month}月"
            if ws_title in wb:
                ws = wb[ws_title]
                for i in range(10,41):
                    ws[f"M{i}"] = f"=IF(AND(COUNTA(H{i})=1,COUNTA(D{i}:H{i})=4),10, IF(AND(COUNTA(H{i})=1,COUNTA(D{i}:H{i})=5),40, 0))"
                ws["H5"].value = tutor.fullname
                ws["K41"].value = f"=ROUNDDOWN(SUM(K10:K40)/60*{tutor.pay_class},0)"
                ws["L41"].value = f"=ROUNDDOWN(SUM(L10:L40)/60*{tutor.pay_officework},0)"
                ws["M41"].value = f"=ROUNDDOWN(SUM(M10:M40)/240*{tutor.pay_officework}, 0)"
                ws["N41"].value = f"=ROUNDDOWN(SUM(N10:N40)/240*{tutor.pay_officework}, 0)"
                ws["P41"].value = f'=COUNTIF(P10:P40,"○")*{tutor.trans_fee}'
            else:
                ws = wb.copy_worksheet(wb.worksheets[0])
                ws.title = ws_title
            ws["P2"].value = gensen_s
            ws["F2"] = year
            ws["H2"] = month % 12 + 1
            for day in range(31):
                for class_time in range(5):
                    if tutor.class_work[day] & (1 << class_time):
                        ws.cell(row=10+day, column=4+class_time).value = class_length
                        tutor.office_work[day] += officetime_per_class
                ws[f"L{10+day}"].value = tutor.office_work[day]
                ws[f"L{10+day}"].value += tutor.meeting[day]
            wb.save(output_path)
            print(tutor.fullname + f"{year}年{month}月分を出力")
            
    # ---Event Process---
    # 実行
    def exec(self, values):
        if "" in [values["tutor_path"], values["admin_path"], values["template_path"], values["gensen_path"], values["year"], values["month"], values["output_folder"]]:
            print("入力されていない項目があります")
            return
        self.update_config(values)
        if not 1 <= int(values["month"]) <= 12:
            print("月の値が不適切です")
            return
        try:
            password = self.receive_password()
        except PassError:
            return
        except NotFilledError:
            print("パスワードが入力されていません")
            return
        except Exception:
            import traceback
            traceback.print_exc()
            return
        if password != default["password"]:
            print("パスワードが異なります")
            return
        print("処理を実行")
        print("--------------------------処理対象ファイル------------------------------------")
        print("講師情報: " + basename(values["tutor_path"]))
        print("管理票: " + basename(values["admin_path"]))
        print("テンプレート: " + basename(values["template_path"]))
        print("---------------------------------------------------------------------------------------")
        print("処理中...")

        try:
            tutors_params = self.receive_tutors(values["tutor_path"])
        except FileNotFoundError:
            import traceback
            traceback.print_exc()
            print("講師情報が見つかりません。ファイルの場所を確認してください")
            print("指定された場所:" + values["tutor_path"])
            return

        self.init_tutors(tutors_params)

        try:
            classes_params = self.receive_classes(values["admin_path"], int(values["month"]), password)
        except FileNotFoundError:
            import traceback
            traceback.print_exc()
            print("管理表が見つかりません。ファイルの場所を確認してください")
            print("指定された場所:" + values["admin_path"])
            return
        except UnboundLocalError:
            import traceback
            traceback.print_exc()
            print("管理表に問題がある可能性があります")
            print("管理表のパスワードを設定しているか確認してください")
            return
        except:
            print("不明なエラー")
            import traceback
            traceback.print_exc()
            return

        self.set_worktime(classes_params)
        self.set_meetings()

        try:
            gensen_s = self.gensen(values["gensen_path"])
        except FileNotFoundError:
            import traceback
            traceback.print_exc()
            print("源泉徴収税額表が見つかりません。ファイルの場所を確認してください")
            print("指定された場所:" + values["gensen_path"])
            return
        except:
            print("不明なエラー")
            import traceback
            traceback.print_exc()
            return

        try:
            self.make_payslip(values["template_path"], gensen_s, int(values["year"]), int(values["month"]), values["output_folder"])
        except PermissionError:
            import traceback
            traceback.print_exc()
            print("講師の給与明細Excelファイルが開いている可能性があります")
            return
        except FileNotFoundError:
            import traceback
            traceback.print_exc()
            print("給与明細テンプレートが見つかりません。ファイルの場所を確認してください")
            print("指定された場所:" + values["template_path"])
            return

        print("処理を終了しました")

    # ミーティングの追加
    def meeting_setting(self, values):
        tutor_path = values["tutor_path"]
        if tutor_path == None:
            print("講師情報が指定されていません")
            return
        try:
            tutors_params = self.receive_tutors(tutor_path)
        except FileNotFoundError:
            import traceback
            traceback.print_exc()
            print("管理票が見つかりません。ファイルの場所を確認してください")
            print("指定された場所:" + values["admin_path"])
            return
        self.init_tutors(tutors_params)
        tutor_names = self.get_tutor_names()
        try:
            meeting_day, meeting_length, participants = self.receive_meeting(tutor_names)
        except PassError:
            return
        except NotFilledError:
            print("入力されていない項目があります")
            return
        except InvalidValue:
            print("日の値が不適切です")
            return
        except Exception:
            print("不明なエラー")
            import traceback
            traceback.print_exc()
            return
        self.add_meeting(int(meeting_day) - 1, int(meeting_length), participants)
        print("--------------------------ミーティングを追加-----------------------------")
        print("実施日: " + meeting_day + " 日")
        print("実施時間: " + meeting_length + " 分")
        print("参加者: " + ', '.join(participants))
        print("------------------------------------------------------------------------------------")

    # ミーティング一覧を表示
    def list_meeting(self, values):
        print("-----------------------------ミーティング一覧-----------------------------")
        for meeting in self.meetings:
            print(f"実施日: {meeting.meeting_day + 1} 日")
            print(f"実施時間: {meeting.meeting_length} 分")
            print("参加者: " + ', '.join(meeting.participants))
            print("------------------------------------------------------------------------------------")
        if self.meetings == []:
            print("------------------------------------------------------------------------------------")

    # ミーティング情報を削除
    def del_meeting(self, values):
        self.meetings = []
        print("ミーティングをすべて削除しました")
import openpyxl
import msoffcrypto
import tempfile
from datetime import datetime, timedelta
import PySimpleGUI as sg
from os.path import basename
import configparser
import warnings

#warnings.simplefilter("ignore")

def main():
    # コンフィグ読み込み(config.iniを実行ファイルと同フォルダに置く)
    config_name = "config.ini"
    config_ini = configparser.ConfigParser()
    config_ini.read(config_name, encoding="utf-8")
    default = config_ini["DEFAULT"]
    keys = ["tutor_path", "admin_path", "template_path", "year", "output_folder"]

    # イベントループ
    while True:
        event, values = window.read()
        
        if event is None:
            print('終了')
            break

        if event == "実行":
            if "" in [values["tutor_path"], values["admin_path"], values["template_path"], values["year"], values["month"], values["output_folder"]]:
                print("入力されていない項目があります")
            else:
                if default.get("password") != sg.popup_get_text(message="パスワードを入力してください", password_char="*", font=(font_name, font_size)):
                    print("パスワードが間違っています")
                else:
                    for key in keys:
                        config_ini.set("DEFAULT", key, values[key])
                    with open(config_name, "w", encoding="utf-8") as f:
                        config_ini.write(f)
                    print("処理を実行")
                    print("処理対象ファイル:" + basename(values["tutor_path"]) + ", " + basename(values["admin_path"]) + ", " + basename(values["template_path"]))
                    exec(values["tutor_path"], values["admin_path"], values["template_path"], int(values["year"]), int(values["month"]), values["output_folder"], default.get("password"))
                    print("処理を終了しました")

    # ウィンドウの破棄と終了

############################################################################

def excel_date(num):
    return(datetime(1899, 12, 30) + timedelta(days=num))

def exec(path_tutor, path_admin, path_template, year, month, output_folder, password):
    # 講師情報変更時のみでOK
    # 講師情報を取得


    ###################################################################################################

    # 管理票を取得
    with open(path_admin, "rb") as f, tempfile.TemporaryFile() as tf:
        ms_file = msoffcrypto.OfficeFile(f)
        ms_file.load_key(password=password)
        ms_file.decrypt(tf)
        wb = openpyxl.load_workbook(tf, data_only=True)

    # 管理票から対象シートを取得
    trans_table = str.maketrans({"０":"0", "１":"1", "２":"2", "３":"3", "４":"4", "５":"5", "６":"6", "７":"7", "８":"8", "９":"9"})
    sheetnames = []
    for sheetname in wb.sheetnames:
        fixed_name = sheetname.translate(trans_table)
        idx = fixed_name.find("月")
        if idx == -1:
            continue
        if fixed_name[:idx] == str(month):
            sheetnames.append(sheetname)

    # 対象シートから出勤日を取得
    class_times = {"2:00-3:20":0, "3:30-4:50":1, "5:00-6:20":2, "6:30-7:50":3, "8:00-9:20":4}
    for sheetname in sheetnames:
        ws = wb[sheetname]
        MAX_ROW = 100
        MAX_COLUMN = 30
        for i in range(1, MAX_ROW):
            head_cell = ws.cell(row=i, column=2).value
            if (type(head_cell) == int) or (head_cell == None):
                if type(head_cell) == int:
                    date = excel_date(head_cell)
                    #############################
                    # 異なる月の日付があったら警告#
                    #############################
                    if date.month != month:
                        break
                    day = date.day - 1 # 0 ~ 30
                table = ws[f"B{i}":f"BA{i+2}"]
                # シート終了の判定
                if table[2][0].value == None: 
                    break
                class_time = class_times[table[2][0].value]
                for j in range(3, MAX_COLUMN):
                    # 登録されていない講師があったらエラー
                    name = table[0][j].value
                    class1 = table[1][j].value
                    class2 = table[2][j].value
                    if name != None and (class1 != None or class2 != None):
                        ##################
                        # 例外コマの処理  #
                        ##################

                        try: 
                            tutors[name]["勤務"][day] |= (1 << class_time)
                        except:
                            print("存在しない講師です")

    ###################################################################################

    # 給与テンプレートを取得
    wb_template = openpyxl.load_workbook(path_template)
    ws_template = wb_template.worksheets[0]
    ws_template["F2"] = year
    ws_template["H2"] = month

    for name in tutor_names:
        if name == None:
            continue
        tutor = tutors[name]
        fullname = tutor["フルネーム"]
        ws_template["H5"].value = fullname
        ws_template["K41"].value = f"=ROUNDDOWN(SUM(K10:K40)/60*{tutor['授業給']},0)"
        ws_template["R41"].value = f'=COUNTIF(R10:R40,"○")*{tutor["交通費"]}'
        for day in range(31):
            for class_time in range(5):
                ws_template.cell(10 + day, 4 + class_time).value = 80 if tutor["勤務"][day] & (1 << class_time) else None
        wb_template.save(output_folder + "/" + fullname + f"{year}年{month}月.xlsx")
        print(fullname + f"{year}年{month}月.xlsx を出力")
    
if __name__=="__main__":
    main()